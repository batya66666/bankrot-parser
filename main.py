import os
import json
import time
import re
import random
from concurrent.futures import ThreadPoolExecutor, as_completed

from bs4 import BeautifulSoup

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import TimeoutException

# ----------------------------
# Output columns (strict order)
# (Документы сразу после дат)
# ----------------------------
OUTPUT_COLUMNS = [
    "Номер аукциона / лота",
    "Адрес объекта",
    "Начальная цена",
    "Шаг аукциона",
    "Размер задатка",
    "Дата и время начала / окончания торгов",
    "Документы",
    "Статус аукциона",
    "Информация о должнике",
    "Описание объекта",
]

DOCS_SHEET_NAME = "Documents"
MAIN_SHEET_NAME = "ALL"


# ----------------------------
# Persistent store for "already parsed lots"
# ----------------------------
class SeenLotsStore:
    def __init__(self, path: str = "seen_lots.json"):
        self.path = path
        self.seen = set()

    def load(self):
        if not os.path.exists(self.path):
            self.seen = set()
            return self.seen
        try:
            with open(self.path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, list):
                self.seen = set(data)
            elif isinstance(data, dict) and "seen" in data and isinstance(data["seen"], list):
                self.seen = set(data["seen"])
            else:
                self.seen = set()
        except Exception:
            self.seen = set()
        return self.seen

    def add_many(self, urls):
        for u in urls:
            if u:
                self.seen.add(u)

    def save(self):
        tmp = self.path + ".tmp"
        payload = sorted(self.seen)
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        os.replace(tmp, self.path)


# ----------------------------
# Cookies auth helper
# ----------------------------
def load_auth_cookies(cookies_path: str):
    if not cookies_path or not os.path.exists(cookies_path):
        return []
    try:
        with open(cookies_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, list) else []
    except Exception:
        return []


def apply_cookies_to_driver(driver, base_url: str, cookies_path: str):
    cookies = load_auth_cookies(cookies_path)
    if not cookies:
        print("auth_cookies.json не найден или пустой")
        return False

    # ОБЯЗАТЕЛЬНО открыть домен
    driver.get(base_url.rstrip("/") + "/")
    time.sleep(1.2)

    added = 0
    for c in cookies:
        if not isinstance(c, dict):
            continue

        # минимальный набор, без domain/expiry/samesite (самый стабильный вариант)
        cc = {
            "name": c.get("name"),
            "value": c.get("value"),
            "path": c.get("path", "/"),
        }

        if "secure" in c:
            cc["secure"] = bool(c["secure"])
        if "httpOnly" in c:
            cc["httpOnly"] = bool(c["httpOnly"])

        if not cc["name"] or cc["value"] is None:
            continue

        try:
            driver.add_cookie(cc)
            added += 1
        except Exception as e:
            print("cookie skip:", cc.get("name"), e)

    driver.refresh()
    time.sleep(1.2)

    page = driver.page_source.lower()
    is_ok = ("войти" not in page)
    print(f"Cookies applied: {added}, logged_in={is_ok}")

    if not is_ok:
        print("⚠️ Похоже, авторизация не применилось. Обнови auth_cookies.json (bankrotbaza_session + XSRF-TOKEN).")

    return is_ok


# ----------------------------
# Excel helpers (ALL + Documents with hyperlinks)
# ----------------------------
def ensure_workbook(filename: str):
    if os.path.exists(filename):
        wb = load_workbook(filename)
    else:
        wb = Workbook()

    # MAIN sheet
    if MAIN_SHEET_NAME in wb.sheetnames:
        ws_main = wb[MAIN_SHEET_NAME]
    else:
        ws_main = wb.active
        ws_main.title = MAIN_SHEET_NAME
        ws_main.append(OUTPUT_COLUMNS)

    # DOCS sheet
    if DOCS_SHEET_NAME in wb.sheetnames:
        ws_docs = wb[DOCS_SHEET_NAME]
    else:
        ws_docs = wb.create_sheet(DOCS_SHEET_NAME)
        ws_docs.append(["Номер аукциона / лота", "Документы"])

    header_font = Font(bold=True)
    for cell in ws_main[1]:
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
    for cell in ws_docs[1]:
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    ws_main.freeze_panes = "A2"
    ws_docs.freeze_panes = "A2"

    return wb, ws_main, ws_docs


def set_column_widths(ws):
    col_widths = {
        "Номер аукциона / лота": 25,
        "Адрес объекта": 65,
        "Начальная цена": 25,
        "Шаг аукциона": 25,
        "Размер задатка": 25,
        "Дата и время начала / окончания торгов": 45,
        "Документы": 18,
        "Статус аукциона": 25,
        "Информация о должнике": 65,
        "Описание объекта": 100,
    }

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    for idx, name in enumerate(headers, start=1):
        if name in col_widths:
            ws.column_dimensions[get_column_letter(idx)].width = col_widths[name]


def read_existing_lot_keys(ws_docs) -> set:
    """Чтобы не дублировать строки лота на Documents при повторном запуске"""
    keys = set()
    for r in range(2, ws_docs.max_row + 1):
        v = ws_docs.cell(row=r, column=1).value
        if v:
            keys.add(str(v).strip())
    return keys


def append_documents_row_wide(ws_docs, lot_key: str, docs: list) -> int:
    """
    Добавляет 1 строку:
      A = lot_key
      B.. = документы (кликабельные гиперссылки)
    Возвращает номер строки, куда записали.
    """
    r = ws_docs.max_row + 1
    ws_docs.cell(row=r, column=1).value = lot_key

    col = 2
    for name, url in docs:
        if not name or not url:
            continue

        cell = ws_docs.cell(row=r, column=col)
        cell.value = name
        cell.hyperlink = url
        cell.style = "Hyperlink"
        col += 1

    return r



def append_rows_with_documents(rows: list, filename: str):
    """
    rows: list of dict with OUTPUT_COLUMNS + internal field:
      - __docs: list[(name,url)]
    Writes:
      - sheet ALL: основные данные + гиперссылка на строку лота в Documents
      - sheet Documents: 1 строка на 1 лот, документы в ширину (B..)
    """
    if not rows:
        print("Нет новых данных для записи в Excel.")
        return

    wb, ws_main, ws_docs = ensure_workbook(filename)

    # чтобы не добавлять повторные строки на Documents
    existing_lot_keys_docs = read_existing_lot_keys(ws_docs)

    # Map header -> column index in MAIN
    main_headers = [ws_main.cell(row=1, column=c).value for c in range(1, ws_main.max_column + 1)]
    header_to_col = {h: i + 1 for i, h in enumerate(main_headers)}

    docs_col = header_to_col.get("Документы")
    if not docs_col:
        ws_main.cell(row=1, column=ws_main.max_column + 1).value = "Документы"
        docs_col = ws_main.max_column
        header_to_col["Документы"] = docs_col

    added_rows = 0

    for row in rows:
        docs = row.pop("__docs", [])
        out = {c: row.get(c, "") for c in OUTPUT_COLUMNS}

        # 1) write MAIN row
        main_row_index = ws_main.max_row + 1
        for col_name in OUTPUT_COLUMNS:
            col_idx = header_to_col.get(col_name)
            if not col_idx:
                col_idx = ws_main.max_column + 1
                ws_main.cell(row=1, column=col_idx).value = col_name
                header_to_col[col_name] = col_idx
            ws_main.cell(row=main_row_index, column=col_idx).value = out.get(col_name, "")

        # 2) write Documents row "wide" once per lot_key
        lot_key = out.get("Номер аукциона / лота", "").strip()
        docs_count = len([d for d in docs if d and d[1]])

        c_docs = ws_main.cell(row=main_row_index, column=docs_col)

        if not lot_key or docs_count == 0:
            c_docs.value = "Документы (0)"
        else:
            # если уже есть строка на Documents — просто найдём её номер (скан, но быстро)
            if lot_key in existing_lot_keys_docs:
                # ищем строку
                doc_row = 0
                for r in range(2, ws_docs.max_row + 1):
                    v = ws_docs.cell(row=r, column=1).value
                    if v and str(v).strip() == lot_key:
                        doc_row = r
                        break
            else:
                doc_row = append_documents_row_wide(ws_docs, lot_key, docs)
                existing_lot_keys_docs.add(lot_key)

            c_docs.value = f"Документы ({docs_count})"
            c_docs.hyperlink = f"#'{DOCS_SHEET_NAME}'!A{doc_row}"
            c_docs.style = "Hyperlink"

        added_rows += 1

    # widths
    set_column_widths(ws_main)

    ws_docs.column_dimensions["A"].width = 25
    # для документов B.. пусть будет нормальная ширина
    for col_idx in range(2, min(ws_docs.max_column, 40) + 1):  # ограничим 40 колонок на всякий
        ws_docs.column_dimensions[get_column_letter(col_idx)].width = 45

    wb.save(filename)
    print(f"Добавлено строк в {filename}: {added_rows}")



# ----------------------------
# Parser
# ----------------------------
class BankrotParser:
    def __init__(self, base_url: str, headless: bool = False, page_load_timeout: int = 25, cookies_path: str = None):
        self.base_url = base_url.rstrip("/")
        self.cookies_path = cookies_path

        chrome_options = Options()
        if headless:
            chrome_options.add_argument("--headless=new")

        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_argument("--window-size=1400,900")
        chrome_options.add_argument(
            "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0 Safari/537.36"
        )

        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option("useAutomationExtension", False)

        self.driver = webdriver.Chrome(
            service=Service(ChromeDriverManager().install()),
            options=chrome_options
        )
        self.driver.execute_cdp_cmd(
            "Page.addScriptToEvaluateOnNewDocument",
            {"source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"}
        )
        self.driver.set_page_load_timeout(page_load_timeout)
        self.wait = WebDriverWait(self.driver, 12)

        # Apply cookies
        if self.cookies_path:
            apply_cookies_to_driver(self.driver, self.base_url, self.cookies_path)

    def close(self):
        try:
            self.driver.quit()
        except Exception:
            pass

    def _parse_info_wrapper(self, soup, title_text: str):
        wrappers = soup.select("div.lot-info__wrapper")
        target = title_text.strip().lower()

        for w in wrappers:
            h3 = w.select_one("h3.lot-info__title")
            if not h3:
                continue
            if h3.get_text(" ", strip=True).strip().lower() != target:
                continue

            out = {}
            for item in w.select("div.lot-info__item"):
                sub = item.select_one(".lot-info__subtitle")
                val_el = item.select_one(".lot-info__value")
                if not sub or not val_el:
                    continue
                key = sub.get_text(" ", strip=True).strip()
                val = val_el.get_text(" ", strip=True).strip()
                out[key] = val

            return out

        return {}

    def _extract_lot_and_trade_numbers(self, soup):
        el = soup.select_one("span.lot__help")
        if not el:
            return ("Не найдено", "Не найдено")

        text = el.get_text(" ", strip=True)
        lot_num = "Не найдено"
        trade_num = "Не найдено"

        m = re.search(r"Лот\s*№\s*(\d+)", text, flags=re.IGNORECASE)
        if m:
            lot_num = m.group(1)

        m = re.search(r"торги\s*№\s*(\d+)", text, flags=re.IGNORECASE)
        if m:
            trade_num = m.group(1)

        return (lot_num, trade_num)

    def _extract_status(self, soup):
        el = soup.select_one("span.lot__status")
        return el.get_text(" ", strip=True) if el else "Не найдено"

    def _extract_details_info(self, soup):
        out = {}
        for item in soup.select("div.lot-details-info__item"):
            sub = item.select_one(".lot-details-info__subtitle")
            if not sub:
                continue
            key = sub.get_text(" ", strip=True).strip()

            val_el = item.select_one(".lot-details-info__value")
            if not val_el:
                val_el = item.find(["span", "div", "a"])
            if not val_el:
                continue

            link_num = item.select_one("a[data-number]")
            if link_num and key.strip().upper() in {"ИНН", "ОГРН"}:
                out[key] = link_num.get("data-number") or link_num.get_text(" ", strip=True).strip()
            else:
                out[key] = val_el.get_text(" ", strip=True).strip()

        return out

    def _extract_debtor_inn_contact(self, soup):
        d = self._extract_details_info(soup)
        debtor = (
            d.get("Наименование / ФИО")
            or d.get("Полное наименование")
            or d.get("Наименование")
            or d.get("Сведения о должнике")
            or "Не найдено"
        )
        inn = d.get("ИНН", "Не найдено")
        contact = d.get("Контактное лицо", "Не найдено")
        return debtor, inn, contact

    def _extract_address_from_text(self, text: str):
        if not text:
            return "Не найдено"
        t = " ".join(text.split())

        m = re.search(
            r"(?i)(?:расположен\w*|наход\w*)\s+по\s+адрес[ау]?\s*:\s*(.+?)(?=(?:начальн\w*\s+цен|задаток|кадастров\w*\s+номер|$))",
            t
        )
        if m:
            return m.group(1).strip(" ,.;")

        m = re.search(
            r"(?i)по\s+адрес[ау]?\s*:\s*(.+?)(?=(?:начальн\w*\s+цен|задаток|кадастров\w*\s+номер|$))",
            t
        )
        if m:
            return m.group(1).strip(" ,.;")

        m = re.search(
            r"(?i)местонахождение\s*:\s*(.+?)(?=(?:начальн\w*\s+цен|задаток|$))",
            t
        )
        if m:
            return m.group(1).strip(" ,.;")

        return "Не найдено"

    def _extract_address_from_desc_p(self, desc_p):
        for a in desc_p.select("a"):
            use = a.select_one("use")
            href = use.get("xlink:href", "") if use else ""
            if "icon-location" in href:
                addr = a.get_text(" ", strip=True).strip()
                return addr if addr else "Не найдено"
        return self._extract_address_from_text(desc_p.get_text(" ", strip=True))

    def _extract_description_and_address(self, soup):
        content = soup.select_one("div.lot__content.text-break")
        if not content:
            return ("Не найдено", "Не найдено")

        desc_p = content.select_one('p[itemprop="description"]')
        if desc_p:
            desc_text = desc_p.get_text(" ", strip=True).strip()
            address = self._extract_address_from_desc_p(desc_p)
            return (desc_text if desc_text else "Не найдено", address)

        ps = content.select("p")
        full = "\n".join(p.get_text(" ", strip=True).strip() for p in ps if p.get_text(strip=True))
        addr = self._extract_address_from_text(full)
        return (full if full.strip() else "Не найдено", addr)

    def _extract_documents(self, soup):
        docs = []
        for a in soup.select(".lot-documents__wrapper a.lot-documents__link"):
            href = (a.get("href") or "").strip()
            name = a.get_text(" ", strip=True).strip()
            if href and name:
                docs.append((name, href))

        seen = set()
        uniq = []
        for name, href in docs:
            if href not in seen:
                seen.add(href)
                uniq.append((name, href))
        return uniq

    def get_listing_urls(self, category_url: str, max_lots: int = 300):
        lot_links = set()
        current_page = 1

        while len(lot_links) < max_lots:
            print(f"Листинг: страница {current_page}")
            separator = "&" if "?" in category_url else "?"
            url = f"{category_url}{separator}page={current_page}"

            try:
                self.driver.get(url)
            except Exception:
                pass

            try:
                self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "a[href*='/lot/']")))
                time.sleep(2.0)
            except TimeoutException:
                print("Лоты не появились (возможно страницы закончились или контент не прогрузился).")
                break

            soup = BeautifulSoup(self.driver.page_source, "html.parser")
            links = soup.select("a[href*='/lot/']")
            if not links:
                print("Ссылки на лоты не найдены (конец списка).")
                break

            before = len(lot_links)
            for a in links:
                if len(lot_links) >= max_lots:
                    break
                href = (a.get("href") or "").strip()
                if not href or href.startswith("#"):
                    continue
                full_link = href if href.startswith("http") else self.base_url + href
                lot_links.add(full_link.split("#")[0])

            print(f"  найдено ссылок: {len(links)}, уникальных всего: {len(lot_links)}")

            if len(lot_links) == before:
                print("Новых лотов не добавилось — остановка.")
                break

            time.sleep(random.uniform(1.6, 2.8))
            current_page += 1

        return list(lot_links)

    def parse_lot_page(self, url: str):
        try:
            self.driver.get(url)
            self.wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
            time.sleep(random.uniform(1.1, 1.9))

            soup = BeautifulSoup(self.driver.page_source, "html.parser")

            lot_number, trade_number = self._extract_lot_and_trade_numbers(soup)
            prices = self._parse_info_wrapper(soup, "Цены")
            dates = self._parse_info_wrapper(soup, "Даты торгов")

            start_price = prices.get("Начальная", "Не найдено")
            step = prices.get("Шаг повышения", "Не найдено")
            zadatok = prices.get("Задаток", "Отсутствует") or "Отсутствует"

            accept_from = dates.get("Приём заявок с", "Не найдено")
            accept_to = dates.get("Приём заявок до", "Не найдено")
            trade_period = f"{accept_from} — {accept_to}"

            status = self._extract_status(soup)
            description, address = self._extract_description_and_address(soup)

            debtor, inn_debtor, contact_person = self._extract_debtor_inn_contact(soup)
            debtor_info = f"{debtor}; ИНН: {inn_debtor}"
            if contact_person and contact_person != "Не найдено":
                debtor_info += f"; Контакт: {contact_person}"

            auction_lot = f"{trade_number} / {lot_number}"
            docs = self._extract_documents(soup)

            row = {
                "Номер аукциона / лота": auction_lot,
                "Адрес объекта": address,
                "Начальная цена": start_price,
                "Шаг аукциона": step,
                "Размер задатка": zadatok,
                "Дата и время начала / окончания торгов": trade_period,
                "Документы": "",
                "Статус аукциона": status,
                "Информация о должнике": debtor_info,
                "Описание объекта": description,
                "__docs": docs,
            }

            if start_price == "Не найдено" and address == "Не найдено":
                print(f"--> Пропуск 'пустого' лота (нет данных): {url}")
                return (url, None)

            return (url, row)

        except Exception as e:
            print(f"Ошибка парсинга: {url} -> {e}")
            return (url, None)


# ----------------------------
# Parallel helpers
# ----------------------------
def worker_parse(base_url: str, urls, headless: bool, cookies_path: str):
    parser = BankrotParser(base_url, headless=headless, cookies_path=cookies_path)
    out_rows = []
    out_urls = []
    try:
        for url in urls:
            u, row = parser.parse_lot_page(url)
            if row:
                out_rows.append(row)
                out_urls.append(u)
            time.sleep(random.uniform(0.8, 1.6))
        return out_rows, out_urls
    finally:
        parser.close()


def chunk_list(items, n_chunks):
    n_chunks = max(1, int(n_chunks))
    chunks = [[] for _ in range(n_chunks)]
    for idx, item in enumerate(items):
        chunks[idx % n_chunks].append(item)
    return chunks


# ----------------------------
# Entry point
# ----------------------------
if __name__ == "__main__":
    BASE_URL = "https://bankrotbaza.ru"
    APARTMENTS_URL = "https://bankrotbaza.ru/search?comb=all&category%5B%5D=27&type_auction=on&sort=created_desc"

    MAX_LOTS = 500
    WORKERS = 3
    HEADLESS = False

    SEEN_FILE = "seen_lots.json"
    OUT_XLSX = "bankrot_apartments.xlsx"
    COOKIES_FILE = "auth_cookies.json"

    print("Cookies path:", os.path.abspath(COOKIES_FILE), "exists:", os.path.exists(COOKIES_FILE))
    if not os.path.exists(COOKIES_FILE):
        print("❌ Не найден auth_cookies.json. Создай файл рядом со скриптом.")
        raise SystemExit(1)

    # 1) listing
    listing_parser = BankrotParser(BASE_URL, headless=HEADLESS, cookies_path=COOKIES_FILE)
    try:
        print("Сбор ссылок (квартиры)...")
        all_links = listing_parser.get_listing_urls(APARTMENTS_URL, max_lots=MAX_LOTS)
    finally:
        listing_parser.close()

    print(f"Собрано ссылок: {len(all_links)}")

    # 2) filter seen
    store = SeenLotsStore(SEEN_FILE)
    seen = store.load()

    new_links = [u for u in all_links if u not in seen]
    print(f"Новых (непарсенных) лотов: {len(new_links)} / уже было: {len(all_links) - len(new_links)}")

    if not new_links:
        print("Ничего нового — выходим.")
        raise SystemExit(0)

    # 3) parallel parse
    chunks = chunk_list(new_links, WORKERS)
    results_rows = []
    parsed_urls = []

    print(f"Запуск параллельно: {WORKERS} браузера(ов)")
    with ThreadPoolExecutor(max_workers=WORKERS) as ex:
        futures = []
        for ch in chunks:
            if ch:
                futures.append(ex.submit(worker_parse, BASE_URL, ch, HEADLESS, COOKIES_FILE))

        for f in as_completed(futures):
            part_rows, part_urls = f.result()
            results_rows.extend(part_rows)
            parsed_urls.extend(part_urls)

    print(f"Успешно распарсено: {len(results_rows)}")

    # 4) write excel
    append_rows_with_documents(results_rows, OUT_XLSX)

    # 5) save seen
    store.add_many(parsed_urls)
    store.save()
    print(f"Память сохранена: {SEEN_FILE} (всего: {len(store.seen)})")
